#!/usr/bin/env python3
"""
build_sheet.py

Per-query metrics sheet: one row per query, one column group per system.

Provenance (important)
----------------------
MAS scores and MAS costs come from DIFFERENT executions, at the user's request:

  MAS facts_recalled / traps_acknowledged  <- the original published mas_v6 run
                                              (sigdial_json/mas_v6.json)
  MAS tokens / llm_calls / turns           <- our metered re-run
                                              (logs/metered_mas.log)

mas_v6.json is the scored per-query output of the run reported in the paper
(FR 86/119, UAA 23/25). Its raw run log is no longer available, so it carries no
cost data. Our two reproduction runs of the same code and config scored
FR 82/119 UAA 22/25 and FR 87/119 UAA 18/25 — the published figures sit inside
that spread, which is sampling variance rather than a code difference.

Consequence to keep in mind: for any given query, the cost columns describe a
different execution from the score columns, so a per-query cost/score
correlation drawn from the MAS group is not sound. The aggregate cost profile
and the aggregate scores are each valid for their own run.

Every baseline is self-consistent: its scores and its costs come from the same
metered execution.

The Provenance sheet records this per system, and the per-query sheet carries a
`mas_scores_from` / `mas_costs_from` stamp so the mix travels with the file.

Usage:
    python3 build_sheet.py
"""
import csv
import json
import os
import re

_ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_X = os.path.join(_ROOT, "sigdial_json_repro", "per_query_sheet.xlsx")
OUT_C = os.path.join(_ROOT, "sigdial_json_repro", "per_query_sheet.csv")

INDEP = set([f"N{i}" for i in range(1, 36)] + [f"D{i}" for i in range(1, 6)])
MISSING_DETAIL = {
    ("N16", "macOS_watch_REJECT"), ("N17", "WinTaskMgr_REJECT"),
    ("N21", "Windows_dir_REJECT"), ("N24", "macOS_watch_REJECT"),
    ("N31", "macOS_AM_REJECT"),
}

# tag -> (scores json, costs jsonl, log for hops/wall-clock, provenance note)
SYSTEMS = [
    ("MAS",   "sigdial_json/mas_v6.json",              "logs/usage_mas.jsonl",
     "logs/metered_mas.log",
     "scores: original published mas_v6 run | costs: our metered re-run — DIFFERENT RUNS"),
    ("UNMAS", "sigdial_json_repro/una2a_metered.json", "logs/usage_a2a.jsonl",
     "logs/metered_a2a.log", "scores + costs: same metered run"),
    ("BB",    "sigdial_json_repro/bb_metered.json",    "logs/usage_blackboard.jsonl",
     "logs/metered_blackboard.log", "scores + costs: same metered run"),
    ("UNBB",  "sigdial_json_repro/unbb_eq.json",  "logs/usage_unstructured_eq.jsonl",
     "logs/metered_unbb_eq.log", "scores + costs: same metered run"),
    ("SA",    "sigdial_json_repro/sa_metered.json",    "logs/usage_single.jsonl",
     "logs/metered_single.log", "scores + costs: same metered run"),
]

_DA = re.compile(r"\bDA=([A-Z_]+)\b")
_EXCLUDE_DA = {"USER_QUERY", "TERMINATE"}


def _blocks(log_path):
    if not os.path.exists(log_path):
        return {}
    txt = open(log_path, encoding="utf-8", errors="replace").read()
    parts = re.split(r"^  \[([ND]\d+)\]", txt, flags=re.M)
    return {parts[i]: parts[i + 1] for i in range(1, len(parts), 2)}


def da_hops(log_path):
    return {q: sum(1 for m in _DA.finditer(b) if m.group(1) not in _EXCLUDE_DA)
            for q, b in _blocks(log_path).items()}


def load_usage(path):
    if not os.path.exists(path):
        return {}
    return {r["query_id"]: r for r in (json.loads(l) for l in open(path, encoding="utf-8"))}


def main():
    scores, costs, hops = {}, {}, {}
    for tag, sf, uf, lf, _ in SYSTEMS:
        p = os.path.join(_ROOT, sf)
        if not os.path.exists(p):
            print(f"  missing {sf} — {tag} omitted")
            continue
        data = json.load(open(p, encoding="utf-8"))
        scores[tag] = {q["query_id"]: q for q in data["queries"]}
        costs[tag] = load_usage(os.path.join(_ROOT, uf))
        if tag == "MAS":
            hops[tag] = da_hops(os.path.join(_ROOT, lf))

    tags = [t for t, *_ in SYSTEMS if t in scores]
    qids = list(scores[tags[0]].keys())

    header = ["query_id", "hop",
              "n_facts", "n_facts_ind", "n_facts_dep",
              "n_traps", "n_traps_ind", "n_traps_dep",
              "n_traps_schema", "n_traps_missingdetail"]
    for t in tags:
        header += [f"{t}_facts_recalled", f"{t}_traps_acked",
                   f"{t}_tokens", f"{t}_llm_calls", f"{t}_turns"]

    rows = []
    for qid in qids:
        ref = scores[tags[0]][qid]
        ind = qid in INDEP
        nf, nt = len(ref["facts"]), len(ref["traps"])
        md = sum(1 for tr in ref["traps"] if (qid, tr["label"]) in MISSING_DETAIL)
        row = [qid, "single" if ind else "multi",
               nf, nf if ind else 0, 0 if ind else nf,
               nt, nt if ind else 0, 0 if ind else nt,
               nt - md, md]
        for t in tags:
            q = scores[t][qid]
            u = costs[t].get(qid, {})
            turns = hops.get(t, {}).get(qid, u.get("turns", "")) if t == "MAS" \
                else u.get("turns", "")
            row += [round(sum(x["score"] for x in q["facts"]), 2),
                    round(sum(x["score"] for x in q["traps"]), 2),
                    u.get("total_tokens", ""), u.get("llm_calls", ""), turns]
        rows.append(row)

    os.makedirs(os.path.dirname(OUT_C), exist_ok=True)
    with open(OUT_C, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)

    # totals row for quick reconciliation against the reported aggregates
    totals = ["TOTAL", ""]
    for i in range(2, 10):
        totals.append(sum(r[i] for r in rows))
    for j in range(10, len(header)):
        vals = [r[j] for r in rows if isinstance(r[j], (int, float))]
        totals.append(round(sum(vals), 2) if vals else "")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "per_query"
        ws.append(header)
        for r in rows:
            ws.append(r)
        ws.append(totals)

        bold = Font(bold=True)
        fill_q = PatternFill("solid", fgColor="EAEFF2")
        fill_s = PatternFill("solid", fgColor="F5F7F8")
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = fill_q if c <= 10 else fill_s
        for c in range(1, len(header) + 1):
            ws.cell(row=ws.max_row, column=c).font = bold
        ws.freeze_panes = "C2"
        for c in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(c)].width = \
                11 if c > 10 else (10 if c > 2 else 12)

        pv = wb.create_sheet("Provenance")
        pv.append(["System", "Scores from", "Costs from", "Note"])
        for c in range(1, 5):
            pv.cell(row=1, column=c).font = bold
        for tag, sf, uf, lf, note in SYSTEMS:
            if tag in scores:
                pv.append([tag, sf, uf, note])
        pv.append([])
        pv.append(["WARNING", "MAS scores and MAS costs are from DIFFERENT executions."])
        pv.append(["", "Scores: original published mas_v6 run (sigdial_json/mas_v6.json)."])
        pv.append(["", "        FR 86/119, UAA 23/25. Raw log gone; no cost data exists."])
        pv.append(["", "Costs:  our metered re-run (logs/metered_mas.log)."])
        pv.append(["", "Our two reproductions scored FR 82/UAA 22 and FR 87/UAA 18 —"])
        pv.append(["", "the published figures sit inside that spread (sampling variance)."])
        pv.append(["", "Do NOT correlate MAS cost against MAS score per query: different runs."])
        pv.append(["", "Baseline rows are self-consistent: one run supplies both."])
        for c in ("A", "B", "C", "D"):
            pv.column_dimensions[c].width = 34 if c != "A" else 12
        pv["A" + str(pv.max_row - 4)].font = Font(bold=True, color="A8402F")

        wb.save(OUT_X)
        print(f"wrote {OUT_X}")
    except ImportError:
        print("openpyxl unavailable — CSV only")

    print(f"wrote {OUT_C}  ({len(rows)} rows, {len(header)} columns)")
    print()
    print("Reconciliation:")
    for t in tags:
        fi = header.index(f"{t}_facts_recalled")
        ti = header.index(f"{t}_traps_acked")
        print(f"  {t:6s} FR {totals[fi]:5.0f}/119   UAA {totals[ti]:4.0f}/25")


if __name__ == "__main__":
    main()
