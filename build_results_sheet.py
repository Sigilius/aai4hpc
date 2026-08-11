#!/usr/bin/env python3
"""
build_results_sheet.py

Results workbook for the AAI4HPC paper, rebuilt from the current runs.

Sheets
------
1  FR_UAA          Table II — FR, FR-Single, FR-Multi, FR-M/FR-S, UAA, UAA-Single, UAA-Multi
2  Turns_Cost      Table IV — turn range, mean turns/tokens/calls, mean AND median wall-clock
3  Trap_Subtype    Table III — schema-absence (20) vs missing-detail (5)
4  DA_Distribution Table V   — Structured MAS typed-message composition, full 12-act
                              vocabulary with 0 against acts never emitted
5  per_query       one row per query, every system

Configuration
-------------
Structured MAS is the all-GPT-4o run (logs/metered_mas_4o.log). The shipped MAS
routes ~77% of calls to gpt-4o-mini; the paper states all configurations use
GPT-4o, so the all-4o run is the one that matches the stated control. Every
baseline is unchanged from its own metered run.

All scores are post-fix: score_query() now unwraps display line-wrapping before
phrase matching, which was silently converting correct refusals into misses.
"""
import csv
import json
import os
import re
import statistics as st

_ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_X = os.path.join(_ROOT, "sigdial_json_repro", "results_tables.xlsx")
OUT_D = os.path.join(_ROOT, "sigdial_json_repro")

INDEP = set([f"N{i}" for i in range(1, 36)] + [f"D{i}" for i in range(1, 6)])
MISSING_DETAIL = {
    ("N16", "macOS_watch_REJECT"), ("N17", "WinTaskMgr_REJECT"),
    ("N21", "Windows_dir_REJECT"), ("N24", "macOS_watch_REJECT"),
    ("N31", "macOS_AM_REJECT"),
}

# Full DialogueActType vocabulary, in enum order (core/message_schema.py).
DA_VOCAB = ["USER_QUERY", "REQUEST", "INFORM", "CLARIFY", "DELEGATE", "CAVEAT",
            "VALIDATE", "CHALLENGE", "CONFIRM", "REJECT", "SYNTHESIZE", "TERMINATE"]

SYSTEMS = [
    ("Structured MAS",           "mas_4o.json",         "usage_mas_4o.jsonl",           "metered_mas_4o.log"),
    ("Unstructured MAS",         "una2a_metered.json",  "usage_a2a.jsonl",              "metered_a2a.log"),
    ("Blackboard MAS",           "bb_metered.json",     "usage_blackboard.jsonl",       "metered_blackboard.log"),
    ("Unstr.-Blackboard MAS",    "unbb_eq.json",        "usage_unstructured_eq.jsonl",  "metered_unbb_eq.log"),
    ("Single Agent",             "sa_metered.json",     "usage_single.jsonl",           "metered_single.log"),
]

_DA = re.compile(r"\bDA=([A-Z_]+)\b")
_EXCLUDE_DA = {"USER_QUERY", "TERMINATE"}
_DUR = re.compile(r"^── ANSWER \(([\d.]+)s\) ──", re.M)


def blocks(log):
    p = os.path.join(_ROOT, log)
    if not os.path.exists(p):
        return {}
    txt = open(p, encoding="utf-8", errors="replace").read()
    parts = re.split(r"^  \[([ND]\d+)\]", txt, flags=re.M)
    return {parts[i]: parts[i + 1] for i in range(1, len(parts), 2)}


def wall(log):
    out = {}
    for q, b in blocks(log).items():
        m = _DUR.search(b)
        if m:
            out[q] = float(m.group(1))
    return out


def hops(log):
    """DA-trace hop count per query (Structured MAS only)."""
    return {q: sum(1 for m in _DA.finditer(b) if m.group(1) not in _EXCLUDE_DA)
            for q, b in blocks(log).items()}


def usage(f):
    p = os.path.join(_ROOT, "logs", f)
    if not os.path.exists(p):
        return {}
    return {r["query_id"]: r for r in (json.loads(l) for l in open(p, encoding="utf-8"))}


def pct(c, t):
    return f"{100*c/t:.1f}% ({c:.0f}/{t})" if t else "—"


def main():
    S, U, W, H = {}, {}, {}, {}
    for name, sf, uf, lf in SYSTEMS:
        p = os.path.join(_ROOT, "sigdial_json_repro", sf)
        if not os.path.exists(p):
            print(f"  missing {sf}; skipping {name}")
            continue
        S[name] = {q["query_id"]: q for q in json.load(open(p, encoding="utf-8"))["queries"]}
        U[name] = usage(uf)
        W[name] = wall(os.path.join("logs", lf))
        H[name] = hops(os.path.join("logs", lf))

    names = [n for n, *_ in SYSTEMS if n in S]
    qids = list(S[names[0]].keys())

    # ── Table 1: FR / UAA ────────────────────────────────────────────────────
    t1 = []
    for n in names:
        fs = fm = fsT = fmT = us = um = usT = umT = 0
        for qid, q in S[n].items():
            single = qid in INDEP
            for x in q["facts"]:
                if single: fs += x["score"]; fsT += 1
                else:      fm += x["score"]; fmT += 1
            for t in q["traps"]:
                if single: us += t["score"]; usT += 1
                else:      um += t["score"]; umT += 1
        ratio = (fm / fmT) / (fs / fsT) if fs and fmT else 0
        t1.append([n, pct(fs + fm, fsT + fmT), pct(fs, fsT), pct(fm, fmT),
                   f"{ratio:.2f}", pct(us + um, usT + umT), pct(us, usT), pct(um, umT)])
    h1 = ["System", "FR", "FR-Single", "FR-Multi", "FR-M / FR-S",
          "UAA", "UAA-Single", "UAA-Multi"]

    # ── Table 2: turns and cost ──────────────────────────────────────────────
    t2 = []
    for n in names:
        u = U[n]
        turns = [H[n].get(q, u.get(q, {}).get("turns", 0)) if n == "Structured MAS"
                 else u.get(q, {}).get("turns", 0) for q in qids]
        turns = [t for t in turns if isinstance(t, (int, float))]
        toks = [u[q]["total_tokens"] for q in qids if q in u]
        calls = [u[q]["llm_calls"] for q in qids if q in u]
        w = [W[n][q] for q in qids if q in W[n]]
        t2.append([n, f"{int(min(turns))}–{int(max(turns))}", round(st.mean(turns), 2),
                   round(st.mean(toks)), round(st.mean(calls), 1),
                   round(st.mean(w), 1), round(st.median(w), 1)])
    h2 = ["System", "Turn range", "Mean turns", "Mean tokens", "Mean LLM calls",
          "Mean wall-clock (s)", "Median wall-clock (s)"]

    # ── Table 3: trap subtype ────────────────────────────────────────────────
    t3 = []
    for n in names:
        sa_c = sa_t = md_c = md_t = 0
        for qid, q in S[n].items():
            for t in q["traps"]:
                if (qid, t["label"]) in MISSING_DETAIL:
                    md_c += t["score"]; md_t += 1
                else:
                    sa_c += t["score"]; sa_t += 1
        tot = sa_c + md_c
        t3.append([n, f"{sa_c:.0f} ({100*sa_c/sa_t:.0f}%)", f"{md_c:.0f} ({100*md_c/md_t:.0f}%)",
                   f"{tot:.0f}/25 = {100*tot/25:.0f}%"])
    h3 = ["Config", "Schema-absence (20)", "Missing-detail (5)", "Total"]

    # ── Table 4: DA distribution (Structured MAS) ────────────────────────────
    counts = {d: 0 for d in DA_VOCAB}
    nq = 0
    for q, b in blocks(os.path.join("logs", "metered_mas_4o.log")).items():
        nq += 1
        for m in _DA.finditer(b):
            if m.group(1) in counts:
                counts[m.group(1)] += 1
    total = sum(counts.values())
    t4 = [[d, round(counts[d] / nq, 2), counts[d],
           f"{100*counts[d]/total:.1f}%" if total else "0.0%"] for d in DA_VOCAB]
    t4.sort(key=lambda r: -r[2])
    t4.append(["Total", round(total / nq, 2), total, "100%"])
    h4 = ["Dialogue Act", "Mean / Query", "Total", "% of DAs"]

    # ── Table 5: per query ───────────────────────────────────────────────────
    h5 = ["query_id", "hop", "n_facts", "n_facts_ind", "n_facts_dep",
          "n_traps", "n_traps_ind", "n_traps_dep",
          "n_traps_schema", "n_traps_missingdetail"]
    for n in names:
        tag = {"Structured MAS": "MAS", "Unstructured MAS": "UNMAS",
               "Blackboard MAS": "BB", "Unstr.-Blackboard MAS": "UNBB",
               "Single Agent": "SA"}[n]
        h5 += [f"{tag}_recalled", f"{tag}_acked", f"{tag}_tokens",
               f"{tag}_llm_calls", f"{tag}_turns", f"{tag}_wall_s"]
    t5 = []
    for qid in qids:
        ref = S[names[0]][qid]
        ind = qid in INDEP
        nf, nt = len(ref["facts"]), len(ref["traps"])
        md = sum(1 for t in ref["traps"] if (qid, t["label"]) in MISSING_DETAIL)
        row = [qid, "single" if ind else "multi", nf, nf if ind else 0, 0 if ind else nf,
               nt, nt if ind else 0, 0 if ind else nt, nt - md, md]
        for n in names:
            q = S[n][qid]; u = U[n].get(qid, {})
            turns = H[n].get(qid, u.get("turns", "")) if n == "Structured MAS" else u.get("turns", "")
            row += [round(sum(x["score"] for x in q["facts"]), 2),
                    round(sum(t["score"] for t in q["traps"]), 2),
                    u.get("total_tokens", ""), u.get("llm_calls", ""), turns,
                    W[n].get(qid, "")]
        t5.append(row)

    # ── write ────────────────────────────────────────────────────────────────
    sheets = [("FR_UAA", h1, t1), ("Turns_Cost", h2, t2), ("Trap_Subtype", h3, t3),
              ("DA_Distribution", h4, t4), ("per_query", h5, t5)]
    for nm, h, rows in sheets:
        with open(os.path.join(OUT_D, f"{nm.lower()}.csv"), "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh); w.writerow(h); w.writerows(rows)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook(); wb.remove(wb.active)
        for nm, h, rows in sheets:
            ws = wb.create_sheet(nm)
            ws.append(h)
            for r in rows:
                ws.append(r)
            for c in range(1, len(h) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.fill = PatternFill("solid", fgColor="EAEFF2")
                ws.column_dimensions[get_column_letter(c)].width = 20 if c == 1 else 14
            if nm == "DA_Distribution":
                for c in range(1, len(h) + 1):
                    ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
            ws.freeze_panes = "B2" if nm != "per_query" else "C2"
        wb.save(OUT_X)
        print(f"wrote {OUT_X}")
    except ImportError:
        print("openpyxl unavailable — CSV only")

    for nm, h, rows in sheets:
        print(f"  {nm:16s} {len(rows):3d} rows -> {nm.lower()}.csv")


if __name__ == "__main__":
    main()
